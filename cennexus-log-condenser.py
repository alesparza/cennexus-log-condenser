# standard imports
from openpyxl import Workbook, load_workbook
from tqdm import tqdm
import sys, getopt, csv, os

# defined constants
DEBUG = False # set to False for regular usage
DEBUG_ROWS = 20
DESCRIPTION_COL_NUM = 3
SEND_STRING = 'Send: <STX>2M|'
RECEIVE_STRING = 'Receive: <STX>3O|'
LOGIN_STRING = '2M|1|101'
STORAGE_STRING = '2M|1|103|'
DEFAULT_SAVE = 'parsed.xlsx'
PARSE_STRING = '-parsed-'
DEFAULT_MERGE = 'merged.xlsx'
USAGE = """usage: cennexus-log-condenser.py -i <inputfile> -o <outputfile> 
alt-usage: cennexus-log-condenser.py -d <directory>"""

def convert_csv(filename):
  """Converts a csv to xlsx

  Basically just copies all the rows into a new workbook

  @return a string for the new file (foo.xlsx)
  """
  wb = Workbook()
  ws = wb.active
  with open(filename, 'r') as f:
    for row in csv.reader(f):    
      ws.append(row)
  f.close()
  index = filename.index('.csv')
  filename = filename[:index] + '.xlsx'
  wb.save(filename)
  wb.close()
  return filename


def parse_xlsx(inputfile, outputfile):
  """Parses an xlsx to get only the host Order and Manufacturer messages

  Includes a progress bar.

  Some extra processing to get the date and time separated.

  If the original file was from a .csv, it will delete the temporary .xlsx file.

  @param inputfile the file to process
  @param outputfile the desired name of the output file
  """
  # open the workbooks and get the active sheets
  tqdm.write('Loading workbook: {}'.format(inputfile))
  wb = load_workbook(filename=inputfile, read_only=True)
  ws = wb.active
  nb = Workbook(write_only=True)
  ns = nb.create_sheet()
  row_count = ws.max_row
  tqdm.write('Loaded Workbook!')

  # loop through each row and write it to the new file if it is a valid message
  tqdm.write('Processing data... Please be patient')
  ns.append(['Timestamp', 'Message', 'isReceive', 'isSend',
    'isLogin', 'isStorage', 'Date', 'Hour', 'Minute', 'Time'])

  # for debugging, limit how many rows are processed
  if DEBUG:
    i = 0

  # using a progress bar here
  with tqdm(total=row_count, ascii=True, desc='Working...') as pbar:
    for row in ws.rows:
      if DEBUG:
        i = i + 1
      pbar.update(1)
      message = row[DESCRIPTION_COL_NUM].value
      if message is None: # just in case a row is empty
        continue
      if DEBUG and i < DEBUG_ROWS:
        tqdm.write(message)
        tqdm.write('Starts with SEND?: {}'.format(message.startswith(SEND_STRING)))
        tqdm.write('Starts with RECEIVE?: {}'.format(message.startswith(RECEIVE_STRING)))

      # check for valid messages and process message counts and date/time
      isSend = 0
      isReceive = 0
      isLogin = 0
      isStorage = 0
      if message.startswith(RECEIVE_STRING):
        isReceive = 1
      elif message.startswith(SEND_STRING):
        isSend = 1
        if LOGIN_STRING in message:
          isLogin = 1
        elif STORAGE_STRING in message:
          isStorage = 1
         
      # keep valid messages or move on
      if (isSend == 1 or isReceive == 1):

        # get the various date components
        timestamp = row[0].value
        split_timestamp = timestamp.split(' ')
        date = split_timestamp[0]
        time_string = split_timestamp[1]
        time_array = time_string.split(':')
        hour = int(time_array[0])
        if hour == 12:
          hour = 0
        if split_timestamp[2] == 'PM':
          hour = hour + 12
        minute = int(time_array[1])
        if int(minute) < 10:
          time = str(hour) + ':0' + str(minute) 
        else:
          time = str(hour) + ':' + str(minute)

        # now write the final row
        if DEBUG and i < DEBUG_ROWS:
          tqdm.write('Valid message, appending to new worksheet...')
        ns.append([timestamp, message, isReceive, isSend, 
          isLogin, isStorage, date, hour, minute, time])
        continue
        if DEBUG and i < DEBUG_ROWS:
          tqdm.write('SEND/RECEIVE not found, skipping row')

  # close workbook and progress bar since they are done
  pbar.close()
  wb.close()
  tqdm.write('Data parse complete!')

  # save the new data in a fresh workbook
  tqdm.write('Saving to ' + outputfile + '...')
  nb.save(outputfile)
  nb.close()
  global isCSV
  if isCSV:
    if DEBUG: 
      tqdm.write('Removing temp file {}'.format(str(inputfile)))
    os.remove(inputfile)
    isCSV = False # reset this in case another file is processed later
  tqdm.write('Data saved!')
  return outputfile


def merge_files(dir_source, outputfile=DEFAULT_MERGE):
  """Merges all parsed files into a single file

  @param dir_source the source directory
  """
  files = os.listdir(dir_source)
  file_count = len(files)
  i = 0
  if DEBUG:
    print('Directory contents ({1}): {0}'.format(files, file_count))
  # setup final workbook
  fb = Workbook(write_only=True)
  fs = fb.create_sheet()

  # go through each file, appending rows if it is a parsed file
  with tqdm(total=file_count, ascii=True, desc='Merging files') as pbar:
    for filen in files:
      if PARSE_STRING in filen:
        wb = load_workbook(filename=dir_source + '\\' + filen, read_only=True)
        ws = wb.active
        tqdm.write('Appending {} to master file'.format(filen))
        for row in ws.values:
          fs.append(row)
        wb.close()
        os.remove(dir_source + '\\' + filen)
      else:
        tqdm.write('Skipping {}...'.format(filen))
      pbar.update(1)

  # save the merged file and cleanup
  finalname = dir_source + '\\' + outputfile
  print('Writing merged file...')
  fb.save(finalname)
  fb.close()
  pbar.close()
  print('Merge into {} complete!'.format(finalname))
  return

def process_dir(dir_source):
  """Processes all .xlsx or .csv files in a directory

  @param dir_source the source directory
  """
  files = os.listdir(dir_source)
  file_count = len(files)
  i = 0
  global isCSV
  if DEBUG:
    print('Directory contents ({1}): {0}'.format(files, file_count))

  with tqdm(total=file_count, ascii=True, desc='Processing files...') as pbar:
    for f in files:
      tqdm.write('Next file: {}'.format(f))
      input_file = dir_source + '\\' + f
      # if the input is .csv, convert to .xlsx
      if f.endswith('.csv'):
          tqdm.write('This is a .csv file, converting to .xlsx...')
          input_file = convert_csv(input_file)
          isCSV = True
      parse_xlsx(input_file, dir_source + '\\' + str(i) + '.' +
          str(file_count - 1) + PARSE_STRING + '.xlsx')  
      i = i + 1
      pbar.update(1)
  pbar.close()
  return


# main method
def main(argv):
   """
   This program will read in an .xlsx or .csv and parse out the Order or
   Manufacturer messages.  It also adds some headers for easier pivot tableing.
   Do note the dependencies.
   """
   print('Cennexus Log Host Message Condenser')
   if DEBUG:
      print('Argument count: ' + str(len(argv)))
      print('Arguments: ' + str(argv))

   # set up some defaults
   input_file = ''
   output_file = DEFAULT_SAVE
   dir_source = ''
   global isCSV
   isCSV = False

   # get the arguments for input and output file
   try:
       opts, args = getopt.getopt(argv, 'hi:o:d:',['input=', 'output=', 'dir='])

   except getopt.GetoptError:
      print(USAGE)
      sys.exit(2)

   for opt, arg in opts:
      if DEBUG:
         print('Parsing option {} and argument {}'.format(opt, arg))
      if opt == '-h':
         print(USAGE)
         sys.exit(2)
      elif opt in ('-d', '--dir'):
         dir_source = arg
      elif opt in ('-i', '--input'):
         input_file = arg
      elif opt in ('-o', '--output'):
         output_file = arg

   if DEBUG:
      print('Input file: ' + input_file)
      print('Output file: ' + output_file)
      print('Directory source: ' + dir_source)

   # quit if there is no input file; output file has a default filename
   if input_file == '' and dir_source == '':
      print(USAGE)
      sys.exit(2)

   # handle directory files first
   if dir_source != '':
     if DEBUG: 
       print('Processing files in directory {}'.format(dir_source))
     process_dir(dir_source)
     if output_file != DEFAULT_SAVE:
       merge_files(dir_source, output_file)
     else: 
       merge_files(dir_source)
     print('Bye bye')
     sys.exit(0)

   # otherwise handle the single input file
   else:
      # if the input is .csv, convert to .xlsx
      if input_file.endswith('.csv'):
          print('This is a .csv file, converting to .xlsx...')
          input_file = convert_csv(input_file)
          isCSV = True
      parse_xlsx(input_file, output_file)   
      print('Bye bye')
      sys.exit(0)

# runs main
if __name__ == '__main__':
   main(sys.argv[1:])

